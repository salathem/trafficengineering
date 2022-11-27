from source import Source
from cell import Fundamentaldiagram
import openpyxl as pyxl


class Network:
    def __init__(self, method):
        self.method = method
        self.cells = []
        self.demand = None

    def set_simulation(self, nr_of_steps, delta_time):
        self.nr_of_steps = nr_of_steps
        self.delta_time = delta_time

    def check_cfl(self):
        for cell in self.cells:
            cell.check_cfl()

    def set_scenario(self, fd, scenario, alinea=None, delta_time=None):
        self.fd = fd
        self.scenario = scenario

        if delta_time:
            self.delta_time = delta_time
        else:
            delta_time = self.delta_time

        if self.method == "ctm":
            from cell_ctm import Cell

        elif self.method == "metanet":
            from cell_metanet import Cell

        delta_length = 0.5

        if scenario == 1:
            # initialize all cells
            cells = [
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd, on_ramp_demand=2000),
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd),
            ]
            self.demand = 4000

        elif scenario == 2:
            # initialize all cells
            cells = [
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd, on_ramp_demand=2500),
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd),
            ]
            self.demand = 4000

        elif scenario == 3:
            # initialize all cells
            cells = [
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd, on_ramp_demand=1500),
                Cell(delta_length, 3, delta_time, fd),
                Cell(delta_length, 1, delta_time, fd),
                Cell(delta_length, 3, delta_time, fd),
            ]
            self.demand = 1500

        else:
            # initialize all cells
            cells = [
                Cell(delta_length, 3, fd, delta_time),
                Cell(delta_length, 3, fd, delta_time),
                Cell(delta_length, 3, fd, delta_time),
                Cell(delta_length, 3, fd, delta_time, on_ramp_demand=1500),
                Cell(delta_length, 3, fd, delta_time),
                Cell(delta_length, 3, fd, delta_time),
                Cell(delta_length, 2, fd, delta_time),
                Cell(delta_length, 3, fd, delta_time, on_ramp_demand=4000),
                Cell(delta_length, 2, fd, delta_time),
                Cell(delta_length, 3, fd, delta_time),
                Cell(delta_length, 3, fd, delta_time),
            ]
            self.demand = 3500

        if self.method == "metanet" and alinea:
            for cell in cells:
                if cell.on_ramp_demand:
                    cell.add_on_ramp(alinea)

        # define upstream demand
        demand_upstream_points = [0, 450 / 3600, 3150 / 3600, 3600 / 3600, 5000 / 3600]
        demand_upstream_values = [0, self.demand, self.demand, 0, 0]

        # initialize Network Demand
        self.demand = Source(delta_time, demand_upstream_points, demand_upstream_values)

        self.cells = cells
        self.set_id()
        self.set_neighbours()

    def import_scenario(self, alinea=None):

        wb = pyxl.load_workbook("scenario.xlsx")
        sheet = wb["cells"]

        self.demand = sheet.cell(row=2, column=10).value
        self.delta_time = sheet.cell(row=2, column=11).value

        if self.method == "ctm":
            from cell_ctm import Cell

        elif self.method == "metanet":
            from cell_metanet import Cell

        row = 2
        cells = []
        cell_id = sheet.cell(row=row, column=1).value
        while cell_id:
            cells.append(Cell(sheet.cell(row=row, column=2).value, sheet.cell(row=row, column=3).value, Fundamentaldiagram(sheet.cell(row=row, column=4).value, sheet.cell(row=row, column=5).value, sheet.cell(row=row, column=6).value), self.delta_time, on_ramp_demand=sheet.cell(row=row, column=7).value, beta=sheet.cell(row=row, column=8).value))
            row += 1
            cell_id = sheet.cell(row=row, column=1).value

        if self.method == "metanet" and alinea:

            for cell in cells:
                if cell.on_ramp_demand:
                    cell.add_on_ramp(alinea)

        # define upstream demand
        demand_upstream_points = [0, 450 / 3600, 3150 / 3600, 3600 / 3600, 5000 / 3600]
        demand_upstream_values = [0, self.demand, self.demand, 0, 0]

        # initialize Network Demand
        self.demand = Source(self.delta_time, demand_upstream_points, demand_upstream_values)

        self.cells = cells
        self.set_id()
        self.set_neighbours()

    # link neighbours and network demand
    def set_neighbours(self):
        self.demand.next_cell = self.cells[0]
        self.cells[0].previous_cell = self.demand
        for index in range(len(self.cells)-1):
            self.cells[index].next_cell = self.cells[index+1]
            self.cells[index+1].previous_cell = self.cells[index]

    # set cell id's
    def set_id(self):
        id = 1
        for cell in self.cells:
            cell.id = id
            id += 1





